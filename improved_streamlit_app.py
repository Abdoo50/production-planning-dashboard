import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import warnings
import io
import os
import re
import time
from sit_date_classifier import SITDateClassifier
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


# Suppress warnings
warnings.filterwarnings("ignore")

# Configuration constants from the original scripts
EXCLUDED_RM_CMMF = {
    4300008721, 4300009502, 4300006495, 4300007207, 4300007493, 
    4300000150, 4300000602, 4300010121, 4300000025, 4300006494, 
    7235600692, 4300000411, 4300008722, 4300008766
}

# Set page config
st.set_page_config(
    page_title="Production Planning Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
def initialize_session_state():
    """Initialize session state variables"""
    if 'analysis_completed' not in st.session_state:
        st.session_state.analysis_completed = False
    if 'results' not in st.session_state:
        st.session_state.results = {}
    if 'processing_status' not in st.session_state:
        st.session_state.processing_status = ""
    if 'error_message' not in st.session_state:
        st.session_state.error_message = ""
    if 'sit_summary' not in st.session_state:
        st.session_state.sit_summary = pd.DataFrame()
    if 'target_plan' not in st.session_state:
        st.session_state.target_plan = {}

initialize_session_state()

# Initialize SIT classifier
@st.cache_resource
def get_sit_classifier():
    return SITDateClassifier()

sit_classifier = get_sit_classifier()

# Progress indicator functions
def show_progress(message, progress=None):
    """Show progress message with optional progress bar"""
    st.session_state.processing_status = message
    if progress is not None:
        progress_bar = st.progress(progress)
        return progress_bar
    return None

def clear_progress():
    """Clear progress indicators"""
    st.session_state.processing_status = ""

# Enhanced error handling decorator
def handle_errors(func):
    """Decorator for enhanced error handling"""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            error_msg = f"Error in {func.__name__}: {str(e)}"
            st.session_state.error_message = error_msg
            st.error(error_msg)
            return None
    return wrapper

def _canon_code_str(x) -> str:
    """To a clean text code: strip, and remove trailing .0 from floats read from Excel."""
    s = str(x).strip()
    return re.sub(r"\.0+$", "", s)

def _canon_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.replace(r"\.0+$", "", regex=True)

EXCLUDED_RM_CMMF = {_canon_code_str(x) for x in EXCLUDED_RM_CMMF}


# Title
st.title("📊 Production Planning Dashboard")
st.markdown("---")

# Display processing status
if st.session_state.processing_status:
    st.info(f"🔄 {st.session_state.processing_status}")

# Display error message if any
if st.session_state.error_message:
    st.error(f"❌ {st.session_state.error_message}")
    if st.button("Clear Error"):
        st.session_state.error_message = ""
        st.rerun()

# Sidebar for inputs
st.sidebar.header("📁 File Uploads")

# Production Mode: Use Streamlit file uploaders
demand_planning_file = st.sidebar.file_uploader(
    "Upload Demand Planning File (BOM)", 
    type=["xlsx"], 
    help="Upload the DemandPlanningforF.G-May.xlsx file containing BOM data"
)

sit_files = st.sidebar.file_uploader(
    "Upload SIT Files", 
    type=["xlsx"], 
    accept_multiple_files=True,
    help="Upload SIT_Week1_July_2025.xlsx and other SIT files"
)

# Analysis options
st.sidebar.header("⚙️ Analysis Options")
analysis_type = st.sidebar.radio(
    "Select Analysis Type:",
    ["With Target", "Without Target", "Both"]
)

# NEW: Allocation method selection
st.sidebar.header("🔧 Allocation Method")
allocation_method = st.sidebar.radio(
    "Select Allocation Method:",
    [
        "Equally Distribution",
        "Production Plan Based (Target)",
        "Price Priority Distribution",
    ],
    help=(
        "Choose how to allocate stock among FGs that share the same RM."
        "\n- Equally Distribution: Split stock evenly among FGs."
        "\n- Production Plan Based (Target): Allocate by target requirements (Target × Qty/Unit)."
        "\n- Price Priority Distribution: Allocate by priority of price (selling price or RM value)."
    ),
)

# Date range filtering for SIT (start and end).
st.sidebar.header("📅 Date Range Filter (SIT)")

# Allow the user to select a starting month and year and an ending month and year. When both
# ranges are set to specific values (not "All"), the SIT data will be filtered to include
# only records whose date lies within the inclusive range. If any of the start or end
# selections are "All", no date filtering is applied (consistent with the previous behaviour
# that allowed viewing all months/years). Removing the prior single-month filter ensures that
# filtering is controlled exclusively via this range selector.

range_start_month = st.sidebar.selectbox(
    "Range start month:",
    ["All", "July", "August", "September", "October", "November", "December"],
    help="Select the first month in the range or 'All' to include all months"
)

range_start_year = st.sidebar.selectbox(
    "Range start year:",
    ["All", 2025, 2024, 2026],
    help="Select the first year in the range or 'All' to include all years"
)

range_end_month = st.sidebar.selectbox(
    "Range end month:",
    ["All", "July", "August", "September", "October", "November", "December"],
    help="Select the last month in the range or 'All' to include all months"
)

range_end_year = st.sidebar.selectbox(
    "Range end year:",
    ["All", 2025, 2024, 2026],
    help="Select the last year in the range or 'All' to include all years"
)

# --- Market filter (BOM) ---
st.sidebar.header("🔎 Market Filter (BOM)")
chk_market_imported = st.sidebar.checkbox("Imported", value=True)
chk_market_local = st.sidebar.checkbox("Local", value=True)
chk_market_inhouse = st.sidebar.checkbox("In House", value=True)


# Process button
process_data = st.sidebar.button("🚀 Run Analysis", type="primary")

# Clear results button
# *************************
# 🧪 What-If Simulation Inputs
st.sidebar.header("🧪 What-If Scenario Simulation")

# Only show simulation if analysis is completed
if st.session_state.analysis_completed:
    # ---------- FG target changes (multiple) ----------
    fg_options = list(st.session_state.target_plan.keys())
    fg_selected_multi = st.sidebar.multiselect("🎯 Choose FGs to Modify Target", fg_options)

    fg_new_targets = {}
    for fg in fg_selected_multi:
        default_val = int(st.session_state.target_plan.get(fg, 0))
        fg_new_targets[fg] = st.sidebar.number_input(
            f"New Target for {fg}",
            min_value=0,
            value=default_val,
            step=1,
            key=f"sim_tgt_{fg}"
        )

    # ---------- RM extra SIT changes (multiple) ----------
    rm_options = (
        st.session_state.sit_summary["RM CMMF"].astype(str).unique().tolist()
        if ("sit_summary" in st.session_state and isinstance(st.session_state.sit_summary, pd.DataFrame) and not st.session_state.sit_summary.empty and "RM CMMF" in st.session_state.sit_summary.columns)
        else []
    )
    rm_selected_multi = st.sidebar.multiselect("📦 Choose RMs to Add SIT", rm_options)

    rm_extra_sit = {}
    for rm in rm_selected_multi:
        rm_extra_sit[rm] = st.sidebar.number_input(
            f"Extra SIT for {rm}",
            min_value=0,
            value=0,
            step=1,
            key=f"sim_sit_{rm}"
        )

    simulate = st.sidebar.button("🧪 Run What-If Simulation")
# **************************


# **************************
if st.session_state.analysis_completed:
    if st.sidebar.button("🔄 Clear Results"):
        st.session_state.analysis_completed = False
        st.session_state.results = {}
        st.session_state.error_message = ""
        st.session_state.processing_status = ""
        st.rerun()


def _xl_format_columns(ws, header_to_fmt: dict):
    """Apply number formats by column header name on the first row."""
    header_map = {cell.value: get_column_letter(cell.column) for cell in ws[1] if cell.value}
    for header, num_fmt in header_to_fmt.items():
        col_letter = header_map.get(header)
        if not col_letter:
            continue
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"].number_format = num_fmt

def _xl_highlight_negative(ws, header_name: str, fill: PatternFill = None):
    """Highlight negative numbers ( < 0 ) in a given header column."""
    if fill is None:
        fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")  # Yellow
    header_map = {cell.value: get_column_letter(cell.column) for cell in ws[1] if cell.value}
    col_letter = header_map.get(header_name)
    if not col_letter:
        return
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=fill))


# ==============================================================================
# EXACT FUNCTIONS FROM PASTED_CONTENT SCRIPTS (with enhanced error handling)
# ==============================================================================

@handle_errors
def load_and_clean_bom(file_obj):
    """
    Loads and cleans the BOM sheet. The priority logic for fast runners has been removed.
    """
    try:
        bom_data = pd.read_excel(file_obj, sheet_name="BOM")
        bom_data.columns = bom_data.columns.str.strip()
        st.success(f"Successfully loaded 'BOM' sheet. Shape: {bom_data.shape}")
    except Exception as e:
        error_msg = f"Error loading 'BOM' sheet. Details: {e}"
        raise FileNotFoundError(error_msg)
        
    # NEW critical_cols list with 'Market' included
    critical_cols = ['Family', 'RM CMMF', 'F.G. CMMF', 'RM In Stock', 'Qty / Unit', 'Description', 'RM Description', 'RM Value', 'Market']
    if not all(col in bom_data.columns for col in critical_cols):
        missing = [col for col in critical_cols if col not in bom_data.columns]
        error_msg = f"FATAL ERROR: BOM sheet is missing required columns: {missing}"
        raise ValueError(error_msg)
        
    bom_data.dropna(subset=['Family', 'RM CMMF', 'F.G. CMMF', 'RM In Stock', 'Qty / Unit'], inplace=True)
    for col in ['RM In Stock', 'Qty / Unit', 'RM Value']:
        bom_data[col] = pd.to_numeric(bom_data[col], errors='coerce').fillna(0)

    # If a Target column exists, convert to numeric and drop rows where Target is missing or <= 0.
    if 'Target' in bom_data.columns:
        bom_data['Target'] = pd.to_numeric(bom_data['Target'], errors='coerce').fillna(0)
        # Filter out rows with zero or negative Target
        before_rows = bom_data.shape[0]
        bom_data = bom_data[bom_data['Target'] > 0].copy()
        after_rows = bom_data.shape[0]
        if before_rows != after_rows:
            st.info(f"Dropped {before_rows - after_rows} BOM rows with missing or zero Target values.")

    st.success(f"BOM data cleaned. Working with all {bom_data.shape[0]} rows.")

    # Ensure codes are canonical strings for consistent joins and isin()
    bom_data["RM CMMF"] = _canon_series(bom_data["RM CMMF"])
    bom_data["F.G. CMMF"] = _canon_series(bom_data["F.G. CMMF"])

    return bom_data

# ==============================================================================
# ALLOCATION FUNCTIONS
# These functions perform stock allocation among FGs that share the same RM.
# allocate_equal: distributes stock equally among FGs using the same RM.
# allocate_by_target: distributes stock proportionally to each FG's target requirement (Target × Qty/Unit) with
# a baseline value for FGs whose target is zero, ensuring no FG is starved of stock.
# ==============================================================================

@handle_errors
def allocate_equal(data: pd.DataFrame, total_stock_col: str, suffix: str = "") -> pd.DataFrame:
    """Allocate stock equally among FGs that share the same RM.

    This function calculates the number of FGs sharing each RM and divides the available stock
    equally among them. FGs that are the sole consumer of an RM receive the full amount.

    Args:
        data: DataFrame containing BOM or SIT-enhanced data.
        total_stock_col: Name of the column containing total stock for each RM.
        suffix: Optional suffix for the allocated stock column name (e.g., '_sit' or '_sim').

    Returns:
        DataFrame with an added 'Allocated Stock{suffix}' column containing the allocated stock per FG.
    """
    if data is None or data.empty:
        return data
    df = data.copy()
    allocated_stock_col = f"Allocated Stock{suffix}"
    # Count how many FGs use each RM
    rm_fg_counts = df.groupby('RM CMMF')['F.G. CMMF'].transform('nunique')
    # Initialize allocation with total stock
    df[allocated_stock_col] = df[total_stock_col]
    # For each RM group, distribute equally among FGs that can produce at least one unit
    for rm, group in df.groupby('RM CMMF'):
        indices = group.index
        n = len(group)
        total_stock = group[total_stock_col].iloc[0]
        # average stock per FG if divided equally
        avg_stock = total_stock / n if n > 0 else 0
        # quantity required per FG unit
        qtys = group['Qty / Unit'].astype(float).values
        # determine which FGs can produce at least one unit from equal share
        producible_mask = qtys <= avg_stock
        # assign equal weight only to producible FGs
        weights = np.where(producible_mask, 1.0, 0.0)
        weight_sum = weights.sum()
        if weight_sum > 0:
            # normalised weights for FGs that can produce
            weights_norm = weights / weight_sum
            df.loc[indices, allocated_stock_col] = weights_norm * total_stock
        else:
            # if no FG can produce, fall back to equal share
            df.loc[indices, allocated_stock_col] = total_stock / n if n > 0 else 0
    return df


@handle_errors
def allocate_by_target(
    data: pd.DataFrame,
    total_stock_col: str,
    target_plan: dict,
    suffix: str = ""
) -> pd.DataFrame:
    """Allocate stock among FGs based on their target requirements.

    Stock is allocated proportionally to each FG's target requirement (Target × Qty/Unit). When no
    positive target is defined within a group, the allocation falls back to equal distribution. FGs
    with zero target receive no allocation when positive targets exist in their RM group. If all targets
    in a group are zero or missing, the allocation falls back to equal distribution among FGs.

    Args:
        data: DataFrame containing BOM or SIT-enhanced data.
        total_stock_col: Name of the column containing total stock for each RM.
        target_plan: Dictionary mapping FG codes to target quantities. If empty, the 'Target'
            column in `data` is used (defaulting to 0 when missing).
        suffix: Optional suffix for the allocated stock column name.

    Returns:
        DataFrame with an added 'Allocated Stock{suffix}' column and 'Target Weight' column
        containing the allocation weight per FG.
    """
    if data is None or data.empty:
        return data
    df = data.copy()
    allocated_stock_col = f"Allocated Stock{suffix}"
    # Determine source of target values: use target_plan if provided, otherwise use 'Target' column
    if target_plan:
        df['target_value'] = (
            df['F.G. CMMF']
            .astype(str)
            .str.replace(r"\.0+$", "", regex=True)
            .str.strip()
            .map(target_plan)
            .fillna(0)
        )
    else:
        # Use the 'Target' column directly (fill missing with 0)
        df['target_value'] = pd.to_numeric(df.get('Target', 0), errors='coerce').fillna(0)

    # Initialize allocated stock with total stock (default)
    df[allocated_stock_col] = df[total_stock_col]
    # Process each RM group independently
    for rm, group in df.groupby('RM CMMF'):
        indices = group.index
        n = len(group)
        total_stock = group[total_stock_col].iloc[0]
        targets = group['target_value'].astype(float).values
        qtys = group['Qty / Unit'].astype(float).values
        if n > 1:
            # Compute average stock per FG if divided equally
            avg_stock = total_stock / n if n > 0 else 0
            # Determine FGs that can produce at least one unit and have positive target
            producible_mask = (qtys <= avg_stock) & (targets > 0)
            if producible_mask.any():
                eff_targets = np.where(producible_mask, targets, 0)
                # Compute required RM for each FG: eff_target × qty per unit
                reqs = eff_targets * qtys
                total_req = reqs.sum()
                if total_req > 0:
                    weights = reqs / total_req
                else:
                    weights = np.zeros(n)
            else:
                # If no FGs meet criteria, fall back to equal distribution among producible FGs (qtys <= avg_stock)
                alt_mask = qtys <= avg_stock
                if alt_mask.any():
                    weights = np.where(alt_mask, 1.0 / alt_mask.sum(), 0.0)
                else:
                    # If no FGs can produce at all, distribute equally among all
                    weights = np.ones(n) / n
            # Assign allocated stock and target weight
            df.loc[indices, allocated_stock_col] = weights * total_stock
            df.loc[indices, 'Target Weight'] = weights
        else:
            # Single FG consumes all available stock
            df.loc[indices, allocated_stock_col] = total_stock
            df.loc[indices, 'Target Weight'] = 1.0
    # Drop temporary target_value column
    df = df.drop(columns=['target_value'], errors='ignore')
    return df


@handle_errors
def allocate_by_price(
    data: pd.DataFrame,
    total_stock_col: str,
    suffix: str = ""
) -> pd.DataFrame:
    """Allocate stock among FGs based on price priority.

    Stock is allocated proportionally to each FG's price value.  If a `Price` column exists
    in the BOM, it is used directly.  Otherwise, the raw-material value per FG is used
    (\ `RM Value × Qty / Unit\`).  When all price values in a group are zero or missing,
    allocation falls back to equal distribution.

    Args:
        data: DataFrame containing BOM or SIT-enhanced data.
        total_stock_col: Name of the column containing total stock for each RM.
        suffix: Optional suffix for the allocated stock column name (e.g., '_sit' or '_sim').

    Returns:
        DataFrame with an added 'Allocated Stock{suffix}' column and 'Price Weight' column
        containing the allocation weight per FG.
    """
    if data is None or data.empty:
        return data
    df = data.copy()
    allocated_stock_col = f"Allocated Stock{suffix}"
    # Determine price value per FG: prefer 'Price' column (selling price). If missing, use RM Value × Qty/Unit.
    if 'Price' in df.columns:
        df['price_value'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0)
    else:
        # Use RM Value × Qty / Unit as proxy for price priority
        df['price_value'] = pd.to_numeric(df.get('RM Value', 0), errors='coerce').fillna(0) * df['Qty / Unit']

    # Initialize allocated stock with total stock
    df[allocated_stock_col] = df[total_stock_col]
    # Process each RM group independently
    for rm, group in df.groupby('RM CMMF'):
        indices = group.index
        n = len(group)
        total_stock = group[total_stock_col].iloc[0]
        prices = group['price_value'].astype(float).values
        qtys = group['Qty / Unit'].astype(float).values
        if n > 1:
            # average stock per FG
            avg_stock = total_stock / n if n > 0 else 0
            # Only consider FGs that can produce at least one unit and have positive price
            producible_mask = (qtys <= avg_stock) & (prices > 0)
            if producible_mask.any():
                price_masked = np.where(producible_mask, prices, 0)
                total_price = price_masked.sum()
                if total_price > 0:
                    weights = price_masked / total_price
                else:
                    weights = np.zeros(n)
            else:
                # If no FGs meet criteria, fall back to equal distribution among producible FGs (qtys <= avg_stock)
                alt_mask = qtys <= avg_stock
                if alt_mask.any():
                    weights = np.where(alt_mask, 1.0 / alt_mask.sum(), 0.0)
                else:
                    # If no FGs can produce, equal distribution among all
                    weights = np.ones(n) / n
            df.loc[indices, allocated_stock_col] = weights * total_stock
            df.loc[indices, 'Price Weight'] = weights
        else:
            df.loc[indices, allocated_stock_col] = total_stock
            df.loc[indices, 'Price Weight'] = 1.0
    # Clean up temporary column
    df = df.drop(columns=['price_value'], errors='ignore')
    return df

@handle_errors
def allocate_common_rms(data, total_stock_col, suffix=''):
    """Identifies common RMs and allocates their stock equally among the FGs that use them."""
    # ... [This function remains unchanged] ...
    print(f"--- Allocating stock from '{total_stock_col}' for common RMs ---")
    rm_fg_counts = data.groupby('RM CMMF')['F.G. CMMF'].transform('nunique')
    is_common_mask = rm_fg_counts > 1
    allocated_stock_col = f'Allocated Stock{suffix}'
    data[allocated_stock_col] = data[total_stock_col]
    data.loc[is_common_mask, allocated_stock_col] = (data.loc[is_common_mask, total_stock_col] / rm_fg_counts[is_common_mask])
    print("Stock allocation complete.")
    return data
def run_bottleneck_analysis_without_target(data, total_stock_col, allocated_stock_col, suffix=''):
    """
    Runs the full production potential analysis on a dataframe, with corrected logic
    for calculating required materials for stalled products.
    """
    analysis_data = data.copy()
    fg_units_possible_col = f'FG Units Possible{suffix}'
    max_fg_units_col = f'Max FG Units{suffix}'
    potential_units_col = f'Potential Units (if 0-stock RM procured){suffix}'
    is_bottleneck_col = f'Is Bottleneck Component{suffix}'
    required_rm_col = f'Required RM for Max Production{suffix}'
    surplus_rm_col = f'RM Surplus after Production{suffix}'

    # --- Step 1: Calculate production potential ---
    analysis_data[fg_units_possible_col] = np.divide(
        analysis_data[allocated_stock_col], analysis_data['Qty / Unit'],
        out=np.zeros_like(analysis_data[allocated_stock_col], dtype=float),
        where=(analysis_data['Qty / Unit'] != 0)).astype(int)
    
    analysis_data['temp_units_for_min'] = np.where(
        analysis_data['RM CMMF'].isin(EXCLUDED_RM_CMMF),
        np.inf, analysis_data[fg_units_possible_col])
    
    max_fg_units_values = analysis_data.groupby('F.G. CMMF')['temp_units_for_min'].transform('min')
    analysis_data[max_fg_units_col] = max_fg_units_values.replace([np.inf, -np.inf], 0).astype(int)
    analysis_data.drop(columns=['temp_units_for_min'], inplace=True)
    
    def get_hypothetical_min(series):
        non_zero_units = series[series > 0]
        return non_zero_units.min() if not non_zero_units.empty else 0
    analysis_data[potential_units_col] = analysis_data.groupby('F.G. CMMF')[fg_units_possible_col].transform(get_hypothetical_min)

    # --- Step 2: Create family summary ---
    # Correct logic: sum once per unique FG, then aggregate by Family.
    family_summary = (
        analysis_data[["Family", "F.G. CMMF", max_fg_units_col]]
        .groupby(["Family", "F.G. CMMF"], as_index=False)[max_fg_units_col]
        .first()
        .groupby("Family", as_index=False)[max_fg_units_col]
        .sum()
        .rename(columns={max_fg_units_col: f"Total Producible Units by Family{suffix}"})
    )
    
    # --- Step 3: CORRECTED LOGIC FOR REQUIREMENT & SURPLUS ---
    
    # First, calculate the standard requirement based on ACTUAL production (Max FG Units).
    analysis_data[required_rm_col] = analysis_data[max_fg_units_col] * analysis_data['Qty / Unit']
    
    # Determine which components are the bottleneck
    analysis_data[is_bottleneck_col] = analysis_data[fg_units_possible_col] == analysis_data[max_fg_units_col]
    
    # Identify the specific rows that are bottlenecks for STALLED products (Max FG Units = 0)
    stalled_bottleneck_mask = (analysis_data[is_bottleneck_col] == True) & (analysis_data[max_fg_units_col] == 0)

    # For ONLY these specific rows, overwrite the 'Required RM' to show the amount needed to reach the potential.
    if stalled_bottleneck_mask.any():
        analysis_data.loc[stalled_bottleneck_mask, required_rm_col] = \
            analysis_data.loc[stalled_bottleneck_mask, potential_units_col] * analysis_data.loc[stalled_bottleneck_mask, 'Qty / Unit']
            
    # Now, calculate the surplus. This will correctly show a negative value (the shortage)
    analysis_data[surplus_rm_col] = analysis_data[allocated_stock_col] - analysis_data[required_rm_col]
    
    # --- Step 4: Prepare final output ---
    output_cols = [
        'Family', 'F.G. CMMF', 'Description', 'RM CMMF', 'RM Description', 'Market', 'RM Value', 'Qty / Unit',
        total_stock_col, allocated_stock_col, 
        fg_units_possible_col, max_fg_units_col, potential_units_col,
        is_bottleneck_col, required_rm_col, surplus_rm_col
    ]
    if suffix == '_sit':
        output_cols.insert(output_cols.index(total_stock_col), 'SIT Quantity')
        output_cols.insert(output_cols.index(total_stock_col), 'RM In Stock')

    bottleneck_analysis = analysis_data[output_cols].sort_values(
        by=['Family', 'F.G. CMMF', is_bottleneck_col], ascending=[True, True, False])
    
    return family_summary, bottleneck_analysis, analysis_data

@handle_errors
def run_bottleneck_analysis_with_target(data, total_stock_col, allocated_stock_col, target_plan, suffix=''):
    """
    Runs the full production potential analysis and adds a 'Target Gap Analysis' by
    calculating requirements and shortages against a specific production target.
    """
    analysis_data = data.copy()
    # --- Original Potential Calculation Columns ---
    fg_units_possible_col = f'FG Units Possible{suffix}'
    max_fg_units_col = f'Max FG Units{suffix}'
    potential_units_col = f'Potential Units (if 0-stock RM procured){suffix}'
    is_bottleneck_col = f'Is Bottleneck Component{suffix}'
    required_rm_col = f'Required RM for Max Production{suffix}'
    surplus_rm_col = f'RM Surplus after Production{suffix}'
    
    # --- NEW Target Gap Analysis Columns ---
    target_qty_col = 'Target Production Qty'
    target_req_rm_col = 'Required RM for Target'
    target_surplus_col = 'Surplus/Shortage vs Target'

    # --- Step 1: Calculate Production Potential ---
    analysis_data[fg_units_possible_col] = np.divide(analysis_data[allocated_stock_col], analysis_data['Qty / Unit'], out=np.zeros_like(analysis_data[allocated_stock_col], dtype=float), where=(analysis_data['Qty / Unit'] != 0)).astype(int)
    analysis_data['temp_units_for_min'] = np.where(analysis_data['RM CMMF'].isin(EXCLUDED_RM_CMMF), np.inf, analysis_data[fg_units_possible_col])
    max_fg_units_values = analysis_data.groupby('F.G. CMMF')['temp_units_for_min'].transform('min')
    analysis_data[max_fg_units_col] = max_fg_units_values.replace([np.inf, -np.inf], 0).astype(int)
    analysis_data.drop(columns=['temp_units_for_min'], inplace=True)
    def get_hypothetical_min(series):
        non_zero_units = series[series > 0]
        return non_zero_units.min() if not non_zero_units.empty else 0
    analysis_data[potential_units_col] = analysis_data.groupby('F.G. CMMF')[fg_units_possible_col].transform(get_hypothetical_min)

    # --- Step 2: Create family summary (based on original potential) ---
    # Correct logic: sum once per unique FG, then aggregate by Family.
    family_summary = (
        analysis_data[["Family", "F.G. CMMF", max_fg_units_col]]
        .groupby(["Family", "F.G. CMMF"], as_index=False)[max_fg_units_col]
        .first()
        .groupby("Family", as_index=False)[max_fg_units_col]
        .sum()
        .rename(columns={max_fg_units_col: f"Total Producible Units by Family{suffix}"})
    )
    
    # --- Step 3: Calculate original requirement and surplus ---
    analysis_data[required_rm_col] = analysis_data[max_fg_units_col] * analysis_data['Qty / Unit']
    analysis_data[is_bottleneck_col] = analysis_data[fg_units_possible_col] == analysis_data[max_fg_units_col]
    stalled_bottleneck_mask = (analysis_data[is_bottleneck_col] == True) & (analysis_data[max_fg_units_col] == 0)
    if stalled_bottleneck_mask.any():
        analysis_data.loc[stalled_bottleneck_mask, required_rm_col] = analysis_data.loc[stalled_bottleneck_mask, potential_units_col] * analysis_data.loc[stalled_bottleneck_mask, 'Qty / Unit']
    analysis_data[surplus_rm_col] = analysis_data[allocated_stock_col] - analysis_data[required_rm_col]
    
    # --- Step 4: Calculate Target-Based Requirement and Surplus ---
    # Map the target plan to a new column. FGs not in the plan get 0.
    analysis_data[target_qty_col] = analysis_data['F.G. CMMF'].map(target_plan).fillna(0).astype(int)
    # Calculate the materials needed for the target
    analysis_data[target_req_rm_col] = analysis_data[target_qty_col] * analysis_data['Qty / Unit']
    # Calculate the gap: stock vs. target requirement. This will be the PO signal.
    analysis_data[target_surplus_col] = analysis_data[allocated_stock_col] - analysis_data[target_req_rm_col]

    # --- Step 5    # Prepare final output with all columns
    output_cols = [
        # Original Columns
        'Family', 'F.G. CMMF', 'Description', 'RM CMMF', 'RM Description', 'Market', 'RM Value', 'Qty / Unit',
        total_stock_col, allocated_stock_col, 
        fg_units_possible_col, max_fg_units_col, potential_units_col,
        is_bottleneck_col, 
        # NEW Target Gap Analysis Columns
        target_qty_col, target_req_rm_col, target_surplus_col,
        required_rm_col, surplus_rm_col # Keep these for now, will drop later if needed
    ]
    if suffix == '_sit':
        output_cols.insert(output_cols.index(total_stock_col), 'SIT Quantity')
        output_cols.insert(output_cols.index(total_stock_col), 'RM In Stock')

    bottleneck_analysis = analysis_data[output_cols].sort_values(
        by=['Family', 'F.G. CMMF', is_bottleneck_col], ascending=[True, True, False])

    # Apply specific modifications for 'With Target' analysis
    if suffix == '' or suffix == '_sit': # Applies to both original and SIT-enhanced 'With Target' outputs
        # Drop specific columns
        columns_to_drop = [
            'Required RM for Max Production_sit'
        ]
        
        bottleneck_analysis = bottleneck_analysis.drop(
            columns=[col for col in columns_to_drop if col in bottleneck_analysis.columns],
            errors='ignore'
        )

        # Reorder 'FG Units Possible_sit Value' to be before 'Max FG Units' or 'Max FG Units_sit'
        if 'FG Units Possible_sit Value' in bottleneck_analysis.columns:
            cols = bottleneck_analysis.columns.tolist()
            fg_units_possible_sit_value_col = 'FG Units Possible_sit Value'
            
            # Find the position of 'Max FG Units' or 'Max FG Units_sit'
            target_col_index = -1
            if 'Max FG Units' in cols:
                target_col_index = cols.index('Max FG Units')
            elif 'Max FG Units_sit' in cols:
                target_col_index = cols.index('Max FG Units_sit')

            if target_col_index != -1:
                # Remove the column from its current position
                cols.remove(fg_units_possible_sit_value_col)
                # Insert it before the target column
                cols.insert(target_col_index, fg_units_possible_sit_value_col)
                bottleneck_analysis = bottleneck_analysis[cols]

    return family_summary, bottleneck_analysis, analysis_data

@handle_errors
def create_fg_summary_view(bottleneck_df, allocated_stock_col, suffix=''):
    """
    Creates the high-level summary view for each Finished Good, now including a
    count of bottleneck components.
    """
    if bottleneck_df.empty: 
        return pd.DataFrame()
    
    df = bottleneck_df.copy()

    # Calculate stranded quantities for incomplete set logic
    wasted_potential_units = df[f'FG Units Possible{suffix}'] - df[f'Max FG Units{suffix}']
    df['stranded_qty'] = wasted_potential_units * df['Qty / Unit']
    df['stranded_value'] = df['stranded_qty'] * df['RM Value']
    
    # Calculate value of allocated stock
    df['RM Allocated Stock Value'] = df[allocated_stock_col] * df['RM Value']
    
    # Calculate the count of bottleneck components (where surplus is negative)
    # Determine which surplus column to use. For SIT analyses the expected column
    # 'RM Surplus after Production{suffix}' may sometimes be missing (e.g. when
    # the report is filtered or certain columns are dropped). If it's not found,
    # fall back to the target surplus column if available.
    surplus_candidates = [
        f'RM Surplus after Production{suffix}',
        'RM Surplus after Production'
    ]
    surplus_col = None
    for candidate in surplus_candidates:
        if candidate in df.columns:
            surplus_col = candidate
            break
    # Fall back to target surplus only if no production-surplus column exists
    if surplus_col is None:
        alt_surplus_col = 'Surplus/Shortage vs Target'
        if alt_surplus_col in df.columns:
            surplus_col = alt_surplus_col
        else:
            # no surplus column available; default bottleneck count to zero
            df['bottleneck_count'] = 0

    if 'bottleneck_count' not in df.columns:
        df['bottleneck_count'] = np.where(df[surplus_col] < 0, 1, 0)

    # Define the aggregation dictionary, now including the new bottleneck count
    agg_dict = {
        'Family': ('Family', 'first'), 
        'Description': ('Description', 'first'),
        'RMAT Stock QTY': (allocated_stock_col, 'sum'), 
        'Stock Value EGP': ('RM Allocated Stock Value', 'sum'), 
        'QTY to be produced': (f'Max FG Units{suffix}', 'first'),
        'incomplete set QTY': ('stranded_qty', 'sum'), 
        'incomplete Set Value': ('stranded_value', 'sum'),
        'Bottleneck': ('bottleneck_count', 'sum')
    }
    
    fg_view = df.groupby('F.G. CMMF').agg(**agg_dict).reset_index()
    
    # Add the new 'Bottleneck' column to the final output order
    final_order = ['Family', 'F.G. CMMF', 'Description', 'RMAT Stock QTY', 'Stock Value EGP', 'QTY to be produced', 'Bottleneck', 'incomplete set QTY', 'incomplete Set Value']
    
    return fg_view[final_order]

@handle_errors
def process_all_sit_sources_streamlit(
    sit_files,
    range_start_month,
    range_start_year,
    range_end_month,
    range_end_year,
):
    """
    Processes SIT data sources uploaded via Streamlit and applies optional date filtering
    across a range of months and years.  If both the start and end selections are
    specific values (not "All"), then only rows whose date (via SITDateClassifier)
    fall within the inclusive range are retained. Otherwise, no date filtering is
    applied.

    Parameters
    ----------
    sit_files : list
        List of uploaded SIT file-like objects.
    range_start_month : str
        The starting month name (e.g., "July").  "All" means no lower bound.
    range_start_year : int or str
        The starting year (e.g., 2025).  "All" means no lower bound.
    range_end_month : str
        The ending month name.  "All" means no upper bound.
    range_end_year : int or str
        The ending year.  "All" means no upper bound.

    Returns
    -------
    pd.DataFrame or None
        A DataFrame with columns ["RM CMMF", "SIT Quantity"] summarised across all
        loaded SIT files, or None if no valid data.
    """
    all_sit_data = []

    # Precompute month order map for months included in the UI
    month_order = {
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }

    # Determine if date filtering should be applied.  Only apply if both start and end are
    # specified (not "All"). Otherwise treat as no filter (all months/years).
    apply_range_filter = (
        range_start_month != "All"
        and range_start_year != "All"
        and range_end_month != "All"
        and range_end_year != "All"
    )

    # If applying range filter, compute all (month, year) pairs between start and end (inclusive)
    date_pairs = []
    if apply_range_filter:
        try:
            start_year_int = int(range_start_year)
            end_year_int = int(range_end_year)
            start_month_num = month_order.get(str(range_start_month), None)
            end_month_num = month_order.get(str(range_end_month), None)
            if start_month_num is None or end_month_num is None:
                apply_range_filter = False
            else:
                # Build list of (year, month) tuples within range
                from datetime import date
                start_date = date(start_year_int, start_month_num, 1)
                end_date = date(end_year_int, end_month_num, 1)
                # If start is after end, swap to ensure non-empty range
                if start_date > end_date:
                    start_date, end_date = end_date, start_date
                # Generate all months between start_date and end_date inclusive
                current_year = start_date.year
                current_month = start_date.month
                while (current_year < end_date.year) or (
                    current_year == end_date.year and current_month <= end_date.month
                ):
                    # Append month name and year for each iteration
                    # Map numeric month back to name using month_order reverse mapping
                    month_name = [k for k, v in month_order.items() if v == current_month]
                    if month_name:
                        date_pairs.append((month_name[0], current_year))
                    # Increment month/year
                    if current_month == 12:
                        current_month = 7  # start of UI range (July) after December
                        current_year += 1
                    else:
                        current_month += 1
                        # Skip months not in month_order (January–June) by advancing to July
                        if current_month < 7:
                            current_month = 7
                # End while
        except Exception:
            apply_range_filter = False

    # Process the uploaded SIT files
    multi_file_frames = []
    # Define the key columns we need from each SIT file.  We include both
    # 'Quantity' and 'Remaining Open Quantity' so that we can later choose
    # whichever quantity field is present for each file.  Additional date
    # columns are included for date filtering, but we will not drop duplicates
    # based on them.
    required_cols_multi = [
        'PO number ',
        'Item No.',
        'Quantity',
        'Remaining Open Quantity',
        'Updated ETA port',
        'ETA  port',
    ]

    for i, sit_file_obj in enumerate(sit_files):
        try:
            # Read the file object directly
            sit_df = pd.read_excel(sit_file_obj)

            # If using range filter, build a cumulative DataFrame for all month-year pairs
            if apply_range_filter and date_pairs:
                filtered_frames = []
                for (month_name, year_val) in date_pairs:
                    try:
                        # Use sit_classifier.filter_by_month on a copy of the DataFrame
                        temp_df = sit_classifier.filter_by_month(
                            sit_df.copy(), month_name, year_val
                        )
                        filtered_frames.append(temp_df)
                    except Exception:
                        continue
                if filtered_frames:
                    sit_df = pd.concat(filtered_frames, ignore_index=True)
                else:
                    sit_df = pd.DataFrame(columns=sit_df.columns)
            # Else, if no range filter, do not filter by date (include all rows)

            # Restrict to required columns, adding missing columns filled with NaN
            available_cols = [col for col in required_cols_multi if col in sit_df.columns]
            sit_df = sit_df[available_cols]
            for col in required_cols_multi:
                if col not in sit_df.columns:
                    sit_df[col] = np.nan
            multi_file_frames.append(sit_df[required_cols_multi])
        except Exception as e:
            st.warning(f"Could not process SIT file {i+1}: {e}")

    if multi_file_frames:
        merged_multi = pd.concat(multi_file_frames, ignore_index=True)
        # Determine which quantity column to use: prefer 'Remaining Open Quantity' if present,
        # otherwise fall back to 'Quantity'.  Convert to numeric for summation.
        qty_col = None
        if 'Remaining Open Quantity' in merged_multi.columns:
            qty_col = 'Remaining Open Quantity'
        elif 'Quantity' in merged_multi.columns:
            qty_col = 'Quantity'
        if qty_col is not None:
            merged_multi[qty_col] = pd.to_numeric(merged_multi[qty_col], errors='coerce').fillna(0)
            grouped_multi = merged_multi.groupby('Item No.', as_index=False)[qty_col].sum()
            grouped_multi.rename(
                columns={'Item No.': 'RM CMMF', qty_col: 'SIT Quantity'},
                inplace=True,
            )
            all_sit_data.append(grouped_multi)

    # Merge and Finalize
    if not all_sit_data:
        return None

    # Combine the results from all sources
    final_sit_data = pd.concat(all_sit_data, ignore_index=True)

    # Perform the final groupby and sum
    final_grouped_data = final_sit_data.groupby('RM CMMF', as_index=False)['SIT Quantity'].sum()

    final_grouped_data["RM CMMF"] = _canon_series(final_grouped_data["RM CMMF"])
    return final_grouped_data

@handle_errors
def load_target_plan(file_obj):
    """
    Load target plan from demand planning file, specifically from 'Target' column.
    """
    try:
        # Try different sheet names for target data
        sheet_names = ["BOM", "Demand Planning", "Target", "Plan", "Sheet1"]
        target_plan = {}
        
        for sheet_name in sheet_names:
            try:
                target_df = pd.read_excel(file_obj, sheet_name=sheet_name)

                # Look for 'F.G. CMMF' or 'CMMF' and 'Target'
                fg_col = None
                if "F.G. CMMF" in target_df.columns: fg_col = "F.G. CMMF"
                elif "CMMF" in target_df.columns: fg_col = "CMMF"

                if fg_col and "Target" in target_df.columns:
                    target_plan = target_df.set_index(fg_col)["Target"].to_dict()
                    st.success(f"Loaded target plan from sheet '{sheet_name}' using 'Target' with {len(target_plan)} targets")

                    # Canonicalize keys to match BOM FG codes
                    target_plan = {_canon_code_str(k): int(v) for k, v in target_plan.items() if pd.notna(v)}
                    return target_plan
                
            except Exception as e:
                continue
        
        # If no target plan found, create a default one
        st.warning("No target plan found in demand planning file using 'Target'. Using default targets.")
        return {}
        
    except Exception as e:
        st.warning(f"Could not load target plan: {e}. Using default targets.")
        return {}

@handle_errors
def create_excel_output_with_target(fg_summary, bottleneck_analysis, family_summary, data_view_type="Original"):
    """
    Excel export for 'With Target'.
    - Original: sheet names have no suffix.
    - With SIT / What-If: sheet names end with '_sit'.
    - Hide *_sit/*_sim requirement/surplus columns for non-Original downloads.
    - Apply Excel number formats + conditional formatting.
    """
    output_buffer = io.BytesIO()
    sheet_sfx = "" if data_view_type == "Original" else "_sit"

    INT_FMT = '#,##0'
    EGP_FMT = '#,##0" ج.م.‏"'

    _hide_non_original = {
        "Required RM for Max Production_sit",
        "RM Surplus after Production_sit",
        "Required RM for Max Production_sim",
        "RM Surplus after Production_sim",
    }

    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        # ---- 1) FG View (write first, then format)
        fg_sheet_name = f"Amiras_FG_View{sheet_sfx}"
        fg_summary.to_excel(writer, sheet_name=fg_sheet_name, index=False)
        ws_fg = writer.sheets[fg_sheet_name]

        # number formats in FG view
        _xl_format_columns(ws_fg, {
            "RMAT Stock QTY": INT_FMT,
            "incomplete set QTY": INT_FMT,
            "Stock Value EGP": EGP_FMT,
            "incomplete Set Value": EGP_FMT,
        })

        # ---- 2) Bottleneck Analysis (clean, add helper value cols, cast ints, write, then format)
        ba = bottleneck_analysis.copy()

        # ⬇️ to this:
        if data_view_type == "Original":
            # Hide these two in the Original (With Target) file
            ba = ba.drop(
                columns=["Required RM for Max Production", "RM Surplus after Production"],
                errors="ignore"
            )
        else:
            # Existing behavior for With SIT / What-If
            ba = ba.drop(
                columns=[c for c in _hide_non_original if c in ba.columns],
                errors="ignore"
            )

        # helper value cols (currency)
        if "FG Units Possible_sit" in ba.columns and "RM Value" in ba.columns:
            ba["FG Units Possible_sit Value"] = ba["FG Units Possible_sit"] * ba["RM Value"]
        if "FG Units Possible" in ba.columns and "RM Value" in ba.columns and "FG Units Possible Value" not in ba.columns:
            ba["FG Units Possible Value"] = ba["FG Units Possible"] * ba["RM Value"]

        # Surplus/Shortage vs Target as int (not float)
        if "Surplus/Shortage vs Target" in ba.columns:
            ba["Surplus/Shortage vs Target"] = pd.to_numeric(ba["Surplus/Shortage vs Target"], errors="coerce").fillna(0).astype(int)

        ba_sheet_name = f"Bottleneck_Analysis{sheet_sfx}"
        ba.to_excel(writer, sheet_name=ba_sheet_name, index=False)
        ws_ba = writer.sheets[ba_sheet_name]

        # number formats in Bottleneck Analysis
        header_to_fmt = {
            "FG Units Possible Value": EGP_FMT,
            "FG Units Possible_sit Value": EGP_FMT,
            "Surplus/Shortage vs Target": INT_FMT,
        }
        _xl_format_columns(ws_ba, header_to_fmt)

        # highlight negative target gaps
        _xl_highlight_negative(ws_ba, "Surplus/Shortage vs Target")

        # ---- 3) Family Summary (no special formatting requested)
        fam_sheet_name = f"Family_Summary{sheet_sfx}"
        family_summary.to_excel(writer, sheet_name=fam_sheet_name, index=False)

    return output_buffer.getvalue()



@handle_errors
def create_excel_output_without_target(fg_summary, bottleneck_analysis, family_summary, data_view_type="Original"):
    """
    Excel export for 'Without Target'.
    - Original: sheet names have no suffix.
    - With SIT / What-If: sheet names end with '_sit'.
    - Hide *_sit/*_sim requirement/surplus columns for non-Original downloads.
    - Apply Excel number formats + conditional formatting (if target column exists).
    """
    output_buffer = io.BytesIO()
    sheet_sfx = "" if data_view_type == "Original" else "_sit"

    INT_FMT = '#,##0'
    EGP_FMT = '#,##0" ج.م.‏"'

    _hide_non_original = {
        "Required RM for Max Production_sit",
        "RM Surplus after Production_sit",
        "Required RM for Max Production_sim",
        "RM Surplus after Production_sim",
    }

    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        # ---- 1) FG View
        fg_sheet_name = f"Amiras_FG_View{sheet_sfx}"
        fg_summary.to_excel(writer, sheet_name=fg_sheet_name, index=False)
        ws_fg = writer.sheets[fg_sheet_name]

        _xl_format_columns(ws_fg, {
            "RMAT Stock QTY": INT_FMT,
            "incomplete set QTY": INT_FMT,
            "Stock Value EGP": EGP_FMT,
            "incomplete Set Value": EGP_FMT,
        })

        # ---- 2) Bottleneck Analysis
        ba = bottleneck_analysis.copy()
        if data_view_type != "Original":
            ba = ba.drop(columns=[c for c in _hide_non_original if c in ba.columns], errors="ignore")

        # Add helper value columns here too (so we can format them as currency)
        if "FG Units Possible_sit" in ba.columns and "RM Value" in ba.columns:
            ba["FG Units Possible_sit Value"] = ba["FG Units Possible_sit"] * ba["RM Value"]
        if "FG Units Possible" in ba.columns and "RM Value" in ba.columns and "FG Units Possible Value" not in ba.columns:
            ba["FG Units Possible Value"] = ba["FG Units Possible"] * ba["RM Value"]

        # Cast Surplus/Shortage vs Target to int if it exists (sometimes present via earlier steps)
        if "Surplus/Shortage vs Target" in ba.columns:
            ba["Surplus/Shortage vs Target"] = pd.to_numeric(ba["Surplus/Shortage vs Target"], errors="coerce").fillna(0).astype(int)

        # NEW ➜ Cast RM Surplus after Production (both views) to int
        surplus_cols = [c for c in ["RM Surplus after Production", "RM Surplus after Production_sit"] if c in ba.columns]
        for c in surplus_cols:
            ba[c] = pd.to_numeric(ba[c], errors="coerce").fillna(0).astype(int)


        ba_sheet_name = f"Bottleneck_Analysis{sheet_sfx}"
        ba.to_excel(writer, sheet_name=ba_sheet_name, index=False)
        ws_ba = writer.sheets[ba_sheet_name]

        _xl_format_columns(ws_ba, {
            "FG Units Possible Value": EGP_FMT,
            "FG Units Possible_sit Value": EGP_FMT,
            "Surplus/Shortage vs Target": INT_FMT,
        })

        # NEW ➜ Format & highlight negatives for RM Surplus after Production
        for c in surplus_cols:
            _xl_format_columns(ws_ba, {c: INT_FMT})
            _xl_highlight_negative(ws_ba, c)

        # If the target gap column exists here, also highlight negatives
        _xl_highlight_negative(ws_ba, "Surplus/Shortage vs Target")

        # ---- 3) Family Summary
        fam_sheet_name = f"Family_Summary{sheet_sfx}"
        family_summary.to_excel(writer, sheet_name=fam_sheet_name, index=False)

    return output_buffer.getvalue()




# ==============================================================================
# ENHANCED VISUALIZATION FUNCTIONS WITH STABLE CONTAINERS
# ==============================================================================

def create_enhanced_visualizations(data_original, data_sit, title_suffix="", unique_id=""):
    """
    Enhanced, business-friendly visuals.
    Assumes inputs are the FG-level summaries produced by create_fg_summary_view()
    with columns like:
      ['Family','F.G. CMMF','Description','RMAT Stock QTY','Stock Value EGP',
       'QTY to be produced','Bottleneck','incomplete set QTY','incomplete Set Value']
    This version adds clear, low-cognitive-load charts + robust checks.
    """
    import plotly.express as px
    import plotly.graph_objects as go
    import pandas as pd

    def _has_cols(df, cols):
        return df is not None and (not df.empty) and all(c in df.columns for c in cols)

    def _safe_topn(df, sort_col, n=10, ascending=False):
        if not _has_cols(df, [sort_col]):
            return pd.DataFrame()
        d = df.copy()
        return d.sort_values(sort_col, ascending=ascending).head(n)

    # Utility to build a comparison long df for side-by-side charts
    def _combine_for_compare(df_orig, df_sit, key_cols, metric_col, label_orig="Original", label_sit="With SIT"):
        frames = []
        if _has_cols(df_orig, key_cols + [metric_col]):
            a = df_orig[key_cols + [metric_col]].copy()
            a["View"] = label_orig
            frames.append(a)
        if _has_cols(df_sit, key_cols + [metric_col]):
            b = df_sit[key_cols + [metric_col]].copy()
            b["View"] = label_sit
            frames.append(b)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # --- helper: force codes to text & strip any ".0" tail that comes from Excel floats
    def _codes_to_text(df):
        if df is None or df.empty:
            return df
        for col in ("F.G. CMMF", "RM CMMF"):
            if col in df.columns:
                df[col] = (
                    df[col]
                    .astype(str)
                    .str.replace(r"\.0+$", "", regex=True)
                    .str.strip()
                )
        return df

    # ensure upstream data uses text codes to avoid numeric/SI formatting
    data_original = _codes_to_text(data_original)
    data_sit = _codes_to_text(data_sit)

    viz_container = st.container()

    with viz_container:
        tab_metrics, tab_prod, tab_bneck, tab_fin = st.tabs([
            "📊 Key Metrics & Overview",
            "📈 Production Analysis",
            "⚠️ Bottleneck Analysis",
            "💰 Financial Analysis"
        ])

        # -----------------------------
        # TAB 1 — METRICS & OVERVIEW
        # -----------------------------
        with tab_metrics:
            st.subheader(f"📊 Key Metrics Dashboard {title_suffix}".strip())

            view_toggle_metrics = st.radio(
                "Select Data View:",
                ["Original Data", "With SIT Data"],
                key=f"metrics_toggle_{unique_id}",
                horizontal=True
            )
            selected = data_sit if view_toggle_metrics == "With SIT Data" else data_original

            if _has_cols(selected, ["F.G. CMMF", "Stock Value EGP", "QTY to be produced"]):
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Total Products", int(selected["F.G. CMMF"].nunique()))
                with col2:
                    st.metric("Total Stock Value (EGP)", f"{selected['Stock Value EGP'].sum():,.0f}")
                with col3:
                    st.metric("Total Producible Units", f"{selected['QTY to be produced'].sum():,.0f}")
                with col4:
                    total_b = int(selected.get("Bottleneck", pd.Series([0]*len(selected))).fillna(0).sum())
                    st.metric("Total Bottlenecks", total_b)
                with col5:
                    tv = selected["Stock Value EGP"].sum()
                    tp = selected["F.G. CMMF"].nunique()
                    st.metric("Avg Stock / Product", f"{(tv/tp if tp else 0):,.0f}")

                # Quick compare cards if both exist
                if _has_cols(data_original, ["QTY to be produced", "Stock Value EGP"]) and \
                   _has_cols(data_sit, ["QTY to be produced", "Stock Value EGP"]):
                    st.subheader("📈 Impact of SIT Data")
                    col1, col2, col3 = st.columns(3)

                    with col1:
                        improvement = data_sit["QTY to be produced"].sum() - data_original["QTY to be produced"].sum()
                        st.metric("Production Improvement", f"{improvement:,.0f}", delta=f"{improvement:,.0f}")

                    with col2:
                        orig_bottlenecks = int(data_original["Bottleneck"].fillna(0).sum())
                        sit_bottlenecks  = int(data_sit["Bottleneck"].fillna(0).sum())
                        bottleneck_reduction = orig_bottlenecks - sit_bottlenecks
                        st.metric(
                            "Bottleneck Reduction",
                            bottleneck_reduction,
                            delta=-bottleneck_reduction,
                        )

                    with col3:
                        value_increase = data_sit["Stock Value EGP"].sum() - data_original["Stock Value EGP"].sum()
                        st.metric("Stock Value Increase", f"{value_increase:,.0f}", delta=f"{value_increase:,.0f}")

                st.markdown("### 🔎 Fast Visuals")

                # 1) Top families by producible units (selected view)
                if _has_cols(selected, ["Family", "QTY to be produced"]):
                    fam = selected.groupby("Family", as_index=False)["QTY to be produced"].sum()
                    fam_top = _safe_topn(fam, "QTY to be produced", n=10)
                    fig = px.bar(fam_top, x="Family", y="QTY to be produced", title="Top Families • Producible Units")
                    st.plotly_chart(fig, use_container_width=True)

                # 2) Delta by family (SIT vs Original) if both available
                if _has_cols(data_original, ["Family","QTY to be produced"]) and _has_cols(data_sit, ["Family","QTY to be produced"]):
                    fam_o = data_original.groupby("Family", as_index=False)["QTY to be produced"].sum().rename(columns={"QTY to be produced":"Orig"})
                    fam_s = data_sit.groupby("Family", as_index=False)["QTY to be produced"].sum().rename(columns={"QTY to be produced":"SIT"})
                    fam_delta = fam_o.merge(fam_s, on="Family", how="outer").fillna(0)
                    fam_delta["Delta Units"] = fam_delta["SIT"] - fam_delta["Orig"]
                    fam_delta_top = _safe_topn(fam_delta, "Delta Units", n=10)
                    fig = px.bar(fam_delta_top, x="Family", y="Delta Units", title="SIT Uplift by Family (Top 10)")
                    st.plotly_chart(fig, use_container_width=True)

                # 3) Treemap of Stock Value (selected view) — ensure FG codes are text in path
                if _has_cols(selected, ["Family","F.G. CMMF","Stock Value EGP"]):
                    treemap_df = selected.groupby(["Family","F.G. CMMF"], as_index=False)["Stock Value EGP"].sum()
                    treemap_df = treemap_df.copy()
                    treemap_df["F.G. CMMF"] = treemap_df["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                    fig = px.treemap(
                        treemap_df,
                        path=["Family","F.G. CMMF"],
                        values="Stock Value EGP",
                        title="Stock Value Distribution (Family ▶ FG)"
                    )
                    st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("No data available for Key Metrics visualization.")

        # -----------------------------
        # TAB 2 — PRODUCTION ANALYSIS
        # -----------------------------
        with tab_prod:
            st.subheader(f"📈 Production Analysis {title_suffix}".strip())

            view_toggle = st.radio(
                "Select Data View:",
                ["Original Data", "With SIT Data", "Compare"],
                key=f"production_toggle_{unique_id}",
                horizontal=True
            )

            if view_toggle == "Compare":
                compare_df = _combine_for_compare(
                    data_original, data_sit,
                    key_cols=["Family","F.G. CMMF","Description"],
                    metric_col="QTY to be produced"
                )
                if not compare_df.empty:
                    # Family side-by-side (top 15 by SIT or Orig combined)
                    fam = compare_df.groupby(["Family","View"], as_index=False)["QTY to be produced"].sum()
                    fam_total = fam.groupby("Family", as_index=False)["QTY to be produced"].sum().rename(columns={"QTY to be produced":"Total"})
                    top_fams = fam_total.sort_values("Total", ascending=False).head(15)["Family"]
                    fam_plot = fam[fam["Family"].isin(top_fams)]
                    fig = px.bar(fam_plot, x="Family", y="QTY to be produced", color="View", barmode="group",
                                 title="Producible Units by Family • Original vs SIT (Top 15)")
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Upload both versions to compare.")
            else:
                selected = data_sit if view_toggle == "With SIT Data" else data_original
                if _has_cols(selected, ["Family","F.G. CMMF","QTY to be produced","Stock Value EGP"]):
                    # Top 20 FGs by producible units — FIX: cast FG to text + categorical axis
                    fg_top = selected.sort_values("QTY to be produced", ascending=False).head(20).copy()
                    fg_top["F.G. CMMF"] = fg_top["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                    fig = px.bar(fg_top, x="F.G. CMMF", y="QTY to be produced", color="Family",
                                 hover_data=["Description"], title="Top FG by Producible Units (Top 20)")
                    fig.update_xaxes(type="category")
                    st.plotly_chart(fig, use_container_width=True)

                    # Scatter: Producible vs Stock Value (FG only in hover, OK as text)
                    size_col = "incomplete Set Value" if "incomplete Set Value" in selected.columns else None
                    scatter_df = selected.copy()
                    if "F.G. CMMF" in scatter_df.columns:
                        scatter_df["F.G. CMMF"] = scatter_df["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                    fig = px.scatter(
                        scatter_df,
                        x="QTY to be produced",
                        y="Stock Value EGP",
                        color="Family",
                        size=size_col,
                        hover_data=["F.G. CMMF","Description"] if "Description" in scatter_df.columns else ["F.G. CMMF"],
                        title="Where Value Meets Capacity (FG level)"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No data available for Production Analysis.")

        # -----------------------------
        # TAB 3 — BOTTLENECK ANALYSIS
        # -----------------------------
        with tab_bneck:
            st.subheader(f"⚠️ Bottleneck Analysis {title_suffix}".strip())

            view_toggle_b = st.radio(
                "Select Data View:",
                ["Original Data", "With SIT Data"],
                key=f"bottleneck_toggle_{unique_id}",
                horizontal=True
            )
            selected = data_sit if view_toggle_b == "With SIT Data" else data_original

            if _has_cols(selected, ["Family","F.G. CMMF","Bottleneck"]):
                # Top 20 worst FG by bottleneck count — FIX: cast FG to text + categorical axis
                fg_b = selected.sort_values("Bottleneck", ascending=False).head(20).copy()
                fg_b["F.G. CMMF"] = fg_b["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                fig = px.bar(fg_b, x="F.G. CMMF", y="Bottleneck", color="Family",
                             hover_data=["Description"], title="Worst FG by Bottleneck Count (Top 20)")
                fig.update_xaxes(type="category")
                st.plotly_chart(fig, use_container_width=True)

                # Incomplete set value by family (no FG axis here)
                if "incomplete Set Value" in selected.columns:
                    fam_inc = selected.groupby("Family", as_index=False)["incomplete Set Value"].sum()
                    fam_inc_top = _safe_topn(fam_inc, "incomplete Set Value", n=12)
                    fig = px.bar(fam_inc_top, x="Family", y="incomplete Set Value",
                                 title="Incomplete Set Value by Family (Top 12)")
                    st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("No data available for Bottleneck Analysis.")

        # -----------------------------
        # TAB 4 — FINANCIAL ANALYSIS
        # -----------------------------
        with tab_fin:
            st.subheader(f"💰 Financial Analysis {title_suffix}".strip())

            view_toggle_fin = st.radio(
                "Select Data View:",
                ["Original Data", "With SIT Data"],
                key=f"financial_toggle_{unique_id}",
                horizontal=True
            )
            selected = data_sit if view_toggle_fin == "With SIT Data" else data_original

            if _has_cols(selected, ["Family","F.G. CMMF","Stock Value EGP"]):
                # Treemap: value concentration — ensure FG codes are text in path
                value_df = selected.groupby(["Family","F.G. CMMF"], as_index=False)["Stock Value EGP"].sum().copy()
                value_df["F.G. CMMF"] = value_df["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                fig = px.treemap(
                    value_df,
                    path=["Family","F.G. CMMF"],
                    values="Stock Value EGP",
                    title="Where the Money Is (Family ▶ FG)"
                )
                st.plotly_chart(fig, use_container_width=True)

                # Top 20 value FGs — FIX: cast FG to text + categorical axis
                fg_val = selected.sort_values("Stock Value EGP", ascending=False).head(20).copy()
                fg_val["F.G. CMMF"] = fg_val["F.G. CMMF"].astype(str).str.replace(r"\.0+$", "", regex=True).str.strip()
                fig = px.bar(fg_val, x="F.G. CMMF", y="Stock Value EGP", color="Family",
                             hover_data=["Description"], title="Top FG by Stock Value (Top 20)")
                fig.update_xaxes(type="category")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("No data available for Financial Analysis.")



# ==============================================================================
# MAIN PROCESSING LOGIC WITH SESSION STATE MANAGEMENT
# ==============================================================================

if process_data:
    if not demand_planning_file:
        st.error("Please upload the Demand Planning File (BOM) to proceed.")
    else:
        # Clear previous errors
        st.session_state.error_message = ""
        
        # Create progress container
        progress_container = st.container()
        
        with progress_container:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Step 1: Load BOM data
                status_text.text("🔄 Loading and cleaning BOM data...")
                progress_bar.progress(10)
                
                bom_data = load_and_clean_bom(demand_planning_file)
                if bom_data is None:
                    st.stop()

                # ✅ Apply Market filter to BOM
                if bom_data is not None and "Market" in bom_data.columns:
                    selected_any = chk_market_imported or chk_market_local or chk_market_inhouse
                    selected_all = chk_market_imported and chk_market_local and chk_market_inhouse

                    # Normalize BOM market text once
                    bom_market_norm = bom_data["Market"].astype(str).str.strip().str.lower()

                    # Build selected set (handle In House spelling variants)
                    selected_norm = set()
                    if chk_market_imported:
                        selected_norm.add("imported")
                    if chk_market_local:
                        selected_norm.add("local")
                    if chk_market_inhouse:
                        selected_norm.update(["in house", "inhouse", "in-house"])

                    if selected_any and not selected_all:
                        keep_mask = bom_market_norm.isin(selected_norm)
                        before_rows = bom_data.shape[0]
                        bom_data = bom_data.loc[keep_mask].copy()
                        after_rows = bom_data.shape[0]
                        picked_labels = ", ".join(
                            [lbl for lbl, on in [("Imported", chk_market_imported),
                                                ("Local", chk_market_local),
                                                ("In House", chk_market_inhouse)] if on]
                        )
                        st.info(f"Applied Market filter: **{picked_labels}** · Rows: {after_rows}/{before_rows}")
                    elif not selected_any:
                        st.warning("All Market options are unchecked — keeping all markets (no filter applied).")


                # Step 2: Process SIT files if uploaded
                status_text.text("🔄 Processing SIT files...")
                progress_bar.progress(20)

                sit_summary = pd.DataFrame()

                if sit_files:
                    # Apply date range filter using the selected start/end month/year values.  If any
                    # of the range selectors are "All", the SIT data will not be filtered by
                    # date (include all rows).  The date filtering itself is handled in
                    # process_all_sit_sources_streamlit.
                    sit_summary = process_all_sit_sources_streamlit(
                        sit_files,
                        range_start_month,
                        range_start_year,
                        range_end_month,
                        range_end_year,
                    )

                    if sit_summary is not None and not sit_summary.empty:
                        # Build a human-readable description of the applied date range.  If all
                        # selectors are "All" then report that no filtering was applied.  Otherwise
                        # report the inclusive range.
                        if (
                            range_start_month != "All"
                            and range_start_year != "All"
                            and range_end_month != "All"
                            and range_end_year != "All"
                        ):
                            filter_description = f"{range_start_month} {range_start_year} – {range_end_month} {range_end_year}"
                        else:
                            filter_description = "All months and years (no filtering)"

                        st.success(
                            f"Successfully processed SIT data for {filter_description}: {len(sit_summary)} unique items"
                        )
                    else:
                        st.info("No valid SIT data processed. Continuing with original stock only.")

                # Step 3: Prepare data for analysis
                status_text.text("🔄 Preparing data for analysis...")
                progress_bar.progress(30)
                
                # Original data (without SIT)
                original_data = bom_data.copy()
                original_data["Total Stock"] = original_data["RM In Stock"]
                
                # SIT-enhanced data (with SIT if available)
                if sit_summary is not None and not sit_summary.empty:
                    # Convert RM CMMF to string for proper merging
                    bom_data["RM CMMF"] = bom_data["RM CMMF"].astype(str)
                    sit_summary["RM CMMF"] = sit_summary["RM CMMF"].astype(str)
                    
                    sit_enhanced_data = pd.merge(
                        bom_data, 
                        sit_summary,
                        on="RM CMMF", 
                        how="left"
                    )
                    sit_enhanced_data["SIT Quantity"] = sit_enhanced_data["SIT Quantity"].fillna(0)
                    sit_enhanced_data["Total Stock"] = sit_enhanced_data["RM In Stock"] + sit_enhanced_data["SIT Quantity"]
                else:
                    sit_enhanced_data = original_data.copy()

                # Step 4: Load target plan for "With Target" analysis
                status_text.text("🔄 Loading target plan...")
                progress_bar.progress(40)
                
                target_plan = {}
                if analysis_type in ["With Target", "Both"]:
                    target_plan = load_target_plan(demand_planning_file)
                    if not target_plan:
                        # Create default target plan based on existing FG CMMFs
                        unique_fgs = bom_data["F.G. CMMF"].unique()
                        target_plan = {fg: 165 for fg in unique_fgs}  # Default target of 165 units (matching example)
                        st.info("Using default target of 165 units for all products.")

                # Store in session state
                st.session_state.sit_summary = sit_summary
                st.session_state.target_plan = target_plan

                # Step 5: Run analysis based on selected type
                results = {}

                if analysis_type in ["Without Target", "Both"]:
                    status_text.text("🔄 Running analysis without target...")
                    progress_bar.progress(50)
                    
                    # Choose allocation function based on user-selected method for 'Without Target' analysis
                    # If 'Equally Distribution' is selected, allocate stock equally among FGs sharing an RM.
                    # Otherwise, allocate based on target plan (or Target column) using target requirements.
                    if allocation_method == "Equally Distribution":
                        original_without_target = allocate_equal(
                            original_data.copy(), "Total Stock", ""
                        )
                    elif allocation_method == "Production Plan Based (Target)":
                        original_without_target = allocate_by_target(
                            original_data.copy(), "Total Stock", target_plan, ""
                        )
                    else:  # Price Priority Distribution
                        original_without_target = allocate_by_price(
                            original_data.copy(), "Total Stock", ""
                        )
                    family_summary_orig, bottleneck_analysis_orig, _ = run_bottleneck_analysis_without_target(
                        original_without_target, "Total Stock", "Allocated Stock", ""
                    )
                    fg_summary_orig = create_fg_summary_view(bottleneck_analysis_orig, "Allocated Stock", "")
                    
                    # Process SIT-ENHANCED data
                    sit_suffix = "_sit" if sit_summary is not None and not sit_summary.empty else ""
                    if allocation_method == "Equally Distribution":
                        sit_without_target = allocate_equal(
                            sit_enhanced_data.copy(), "Total Stock", sit_suffix
                        )
                    elif allocation_method == "Production Plan Based (Target)":
                        sit_without_target = allocate_by_target(
                            sit_enhanced_data.copy(), "Total Stock", target_plan, sit_suffix
                        )
                    else:
                        sit_without_target = allocate_by_price(
                            sit_enhanced_data.copy(), "Total Stock", sit_suffix
                        )
                    family_summary_sit, bottleneck_analysis_sit, _ = run_bottleneck_analysis_without_target(
                        sit_without_target, "Total Stock", f"Allocated Stock{sit_suffix}", sit_suffix
                    )
                    fg_summary_sit = create_fg_summary_view(bottleneck_analysis_sit, f"Allocated Stock{sit_suffix}", sit_suffix)
                    
                    # Compute additional KPI metrics (surplus and shortage value) for without-target analysis
                    def _calc_surplus_shortage_metrics(bn_df):
                        try:
                            # Surplus/shortage values are computed from RM Surplus after Production and RM Value
                            col_name = f"RM Surplus after Production{sit_suffix}" if sit_suffix else "RM Surplus after Production"
                            if col_name not in bn_df.columns:
                                # Fall back to base column
                                col_name = "RM Surplus after Production"
                            surplus_series = pd.to_numeric(bn_df[col_name], errors='coerce').fillna(0)
                            value_series = pd.to_numeric(bn_df.get('RM Value', 0), errors='coerce').fillna(0)
                            total_surplus_val = (surplus_series[surplus_series > 0] * value_series[surplus_series > 0]).sum()
                            total_shortage_val = (surplus_series[surplus_series < 0].abs() * value_series[surplus_series < 0]).sum()
                            return {
                                'Total Surplus Value': total_surplus_val,
                                'Total Shortage Value': total_shortage_val,
                            }
                        except Exception:
                            return {
                                'Total Surplus Value': 0,
                                'Total Shortage Value': 0,
                            }

                    orig_metrics = _calc_surplus_shortage_metrics(bottleneck_analysis_orig)
                    sit_metrics = _calc_surplus_shortage_metrics(bottleneck_analysis_sit)

                    # Store results including KPI metrics
                    results['without_target'] = {
                        'original': {
                            'fg': fg_summary_orig,
                            'bottleneck': bottleneck_analysis_orig,
                            'family': family_summary_orig,
                            'metrics': orig_metrics,
                        },
                        'sit': {
                            'fg': fg_summary_sit,
                            'bottleneck': bottleneck_analysis_sit,
                            'family': family_summary_sit,
                            'metrics': sit_metrics,
                        }
                    }

                if analysis_type in ["With Target", "Both"]:
                    status_text.text("🔄 Running analysis with target...")
                    progress_bar.progress(70)
                    
                    # For analyses that use production targets, choose allocation function based on user selection.
                    # Equally Distribution: allocate stock equally; Production Plan Based: allocate by target; Price Priority: allocate by price.
                    if allocation_method == "Equally Distribution":
                        original_with_target = allocate_equal(
                            original_data.copy(), "Total Stock", ""
                        )
                    elif allocation_method == "Production Plan Based (Target)":
                        original_with_target = allocate_by_target(
                            original_data.copy(), "Total Stock", target_plan, ""
                        )
                    else:
                        original_with_target = allocate_by_price(
                            original_data.copy(), "Total Stock", ""
                        )
                    family_summary_orig_tgt, bottleneck_analysis_orig_tgt, _ = run_bottleneck_analysis_with_target(
                        original_with_target, "Total Stock", "Allocated Stock", target_plan, ""
                    )
                    fg_summary_orig_tgt = create_fg_summary_view(bottleneck_analysis_orig_tgt, "Allocated Stock", "")
                    
                    # Process SIT-enhanced data using selected allocation function
                    sit_suffix = "_sit" if sit_summary is not None and not sit_summary.empty else ""
                    if allocation_method == "Equally Distribution":
                        sit_with_target = allocate_equal(
                            sit_enhanced_data.copy(), "Total Stock", sit_suffix
                        )
                    elif allocation_method == "Production Plan Based (Target)":
                        sit_with_target = allocate_by_target(
                            sit_enhanced_data.copy(), "Total Stock", target_plan, sit_suffix
                        )
                    else:
                        sit_with_target = allocate_by_price(
                            sit_enhanced_data.copy(), "Total Stock", sit_suffix
                        )
                    family_summary_sit_tgt, bottleneck_analysis_sit_tgt, _ = run_bottleneck_analysis_with_target(
                        sit_with_target, "Total Stock", f"Allocated Stock{sit_suffix}", target_plan, sit_suffix
                    )
                    fg_summary_sit_tgt = create_fg_summary_view(bottleneck_analysis_sit_tgt, f"Allocated Stock{sit_suffix}", sit_suffix)
                    
                    # Compute KPI metrics for with-target analysis (using Surplus/Shortage vs Target)
                    def _calc_target_metrics(bn_df):
                        try:
                            # Surplus/shortage vs Target multiplied by RM Value
                            col_name = "Surplus/Shortage vs Target"
                            if col_name not in bn_df.columns:
                                return {
                                    'Total Surplus vs Target Value': 0,
                                    'Total Shortage vs Target Value': 0,
                                }
                            surplus_series = pd.to_numeric(bn_df[col_name], errors='coerce').fillna(0)
                            value_series = pd.to_numeric(bn_df.get('RM Value', 0), errors='coerce').fillna(0)
                            total_surplus_val = (surplus_series[surplus_series > 0] * value_series[surplus_series > 0]).sum()
                            total_shortage_val = (surplus_series[surplus_series < 0].abs() * value_series[surplus_series < 0]).sum()
                            return {
                                'Total Surplus vs Target Value': total_surplus_val,
                                'Total Shortage vs Target Value': total_shortage_val,
                            }
                        except Exception:
                            return {
                                'Total Surplus vs Target Value': 0,
                                'Total Shortage vs Target Value': 0,
                            }

                    orig_target_metrics = _calc_target_metrics(bottleneck_analysis_orig_tgt)
                    sit_target_metrics = _calc_target_metrics(bottleneck_analysis_sit_tgt)

                    # Store results including KPI metrics
                    results['with_target'] = {
                        'original': {
                            'fg': fg_summary_orig_tgt,
                            'bottleneck': bottleneck_analysis_orig_tgt,
                            'family': family_summary_orig_tgt,
                            'metrics': orig_target_metrics,
                        },
                        'sit': {
                            'fg': fg_summary_sit_tgt,
                            'bottleneck': bottleneck_analysis_sit_tgt,
                            'family': family_summary_sit_tgt,
                            'metrics': sit_target_metrics,
                        }
                    }

                # Step 6: Store results in session state
                status_text.text("🔄 Finalizing results...")
                progress_bar.progress(90)
                
                st.session_state.results = results
                st.session_state.analysis_completed = True
                
                # Clear progress indicators
                progress_bar.progress(100)
                status_text.text("✅ Analysis completed successfully!")
                time.sleep(1)  # Brief pause to show completion
                
                # Clear progress container
                progress_container.empty()
                clear_progress()
                
                # Inform the user of the allocation method used
                allocation_msg = f"✅ Analysis completed successfully using **{allocation_method}** allocation method!"
                st.success(allocation_msg)
                
            except Exception as e:
                progress_container.empty()
                st.session_state.error_message = f"An error occurred during processing: {e}"
                st.error(st.session_state.error_message)

# Display results if analysis is completed
if st.session_state.analysis_completed and st.session_state.results:
    
    # Create stable results container
    results_container = st.container()
    
    with results_container:
        
        # Display results based on analysis type
        if analysis_type in ["Without Target", "Both"] and 'without_target' in st.session_state.results:
            st.header("📊 Analysis Without Target")

            # Enhanced visualizations
            create_enhanced_visualizations(
                st.session_state.results['without_target']['original']['fg'],
                st.session_state.results['without_target']['sit']['fg'],
                "(Without Target)",
                "without_target",
            )

            # Display additional metrics for surplus and shortage
            st.subheader("📉 Surplus and Shortage Metrics")
            orig_metrics = st.session_state.results['without_target']['original'].get('metrics', {})
            sit_metrics = st.session_state.results['without_target']['sit'].get('metrics', {})
            # Use two columns to display original vs SIT metrics
            mcol1, mcol2 = st.columns(2)
            with mcol1:
                st.metric(
                    "Surplus Value (Original)",
                    f"{orig_metrics.get('Total Surplus Value', 0):,.0f}",
                )
                st.metric(
                    "Shortage Value (Original)",
                    f"{orig_metrics.get('Total Shortage Value', 0):,.0f}",
                )
            with mcol2:
                st.metric(
                    "Surplus Value (With SIT)",
                    f"{sit_metrics.get('Total Surplus Value', 0):,.0f}",
                )
                st.metric(
                    "Shortage Value (With SIT)",
                    f"{sit_metrics.get('Total Shortage Value', 0):,.0f}",
                )

            # Download buttons
            st.subheader("📥 Download Options")
            col1, col2 = st.columns(2)
            with col1:
                # Original data download
                excel_data_orig = create_excel_output_without_target(
                    st.session_state.results['without_target']['original']['fg'],
                    st.session_state.results['without_target']['original']['bottleneck'],
                    st.session_state.results['without_target']['original']['family'],
                    "Original",
                )
                if excel_data_orig:
                    st.download_button(
                        label="📥 Download FG to RMAT (Without Target) - Original Data",
                        data=excel_data_orig,
                        file_name="FGtoRMAT(Imported+Local)Withouttarget_Original.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            with col2:
                # SIT data download
                excel_data_sit = create_excel_output_without_target(
                    st.session_state.results['without_target']['sit']['fg'],
                    st.session_state.results['without_target']['sit']['bottleneck'],
                    st.session_state.results['without_target']['sit']['family'],
                    "With_SIT",
                )
                if excel_data_sit:
                    st.download_button(
                        label="📥 Download FG to RMAT (Without Target) - With SIT Data",
                        data=excel_data_sit,
                        file_name="FGtoRMAT(Imported+Local)Withouttarget_WithSIT.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        if analysis_type in ["With Target", "Both"] and 'with_target' in st.session_state.results:
            st.header("📈 Analysis With Target")
            
            # Enhanced visualizations
            create_enhanced_visualizations(
                st.session_state.results['with_target']['original']['fg'],
                st.session_state.results['with_target']['sit']['fg'],
                "(With Target)",
                "with_target",
            )

            # Display additional metrics for surplus/shortage vs target
            st.subheader("📉 Surplus/Shortage vs Target Metrics")
            orig_tgt_metrics = st.session_state.results['with_target']['original'].get('metrics', {})
            sit_tgt_metrics = st.session_state.results['with_target']['sit'].get('metrics', {})
            mtcol1, mtcol2 = st.columns(2)
            with mtcol1:
                st.metric(
                    "Surplus Value vs Target (Original)",
                    f"{orig_tgt_metrics.get('Total Surplus vs Target Value', 0):,.0f}",
                )
                st.metric(
                    "Shortage Value vs Target (Original)",
                    f"{orig_tgt_metrics.get('Total Shortage vs Target Value', 0):,.0f}",
                )
            with mtcol2:
                st.metric(
                    "Surplus Value vs Target (With SIT)",
                    f"{sit_tgt_metrics.get('Total Surplus vs Target Value', 0):,.0f}",
                )
                st.metric(
                    "Shortage Value vs Target (With SIT)",
                    f"{sit_tgt_metrics.get('Total Shortage vs Target Value', 0):,.0f}",
                )

            # Download buttons
            st.subheader("📥 Download Options")
            col1, col2 = st.columns(2)
            with col1:
                # Original data download
                excel_data_orig_tgt = create_excel_output_with_target(
                    st.session_state.results['with_target']['original']['fg'],
                    st.session_state.results['with_target']['original']['bottleneck'],
                    st.session_state.results['with_target']['original']['family'],
                    "Original",
                )
                if excel_data_orig_tgt:
                    st.download_button(
                        label="📥 Download FG to RMAT (With Target) - Original Data",
                        data=excel_data_orig_tgt,
                        file_name="FGtoRMAT(Imported+Local)Withtarget_Original.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            with col2:
                # SIT data download
                excel_data_sit_tgt = create_excel_output_with_target(
                    st.session_state.results['with_target']['sit']['fg'],
                    st.session_state.results['with_target']['sit']['bottleneck'],
                    st.session_state.results['with_target']['sit']['family'],
                    "With_SIT",
                )
                if excel_data_sit_tgt:
                    st.download_button(
                        label="📥 Download FG to RMAT (With Target) - With SIT Data",
                        data=excel_data_sit_tgt,
                        file_name="FGtoRMAT(Imported+Local)Withtarget_WithSIT.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

# ==============================================================================
# SIMULATION SECTION
# ==============================================================================  
# 🧪 Run What-If Simulation
if 'simulate' in locals() and simulate:
    st.header("🧪 What-If Scenario Results")

    import pandas as pd
    import io

    # ---- Base SIT table from session (fallback to empty if missing) ----
    sit_simulated = (
        st.session_state.sit_summary.copy()
        if "sit_summary" in st.session_state and isinstance(st.session_state.sit_summary, pd.DataFrame)
        else pd.DataFrame(columns=["RM CMMF", "SIT Quantity"])
    )
    if "RM CMMF" in sit_simulated.columns:
        sit_simulated["RM CMMF"] = sit_simulated["RM CMMF"].astype(str)
    if "SIT Quantity" in sit_simulated.columns:
        sit_simulated["SIT Quantity"] = pd.to_numeric(sit_simulated["SIT Quantity"], errors="coerce").fillna(0)

    # ---- ✅ Apply ALL RM adjustments (adds new rows when needed) ----
    if "rm_extra_sit" in locals() and isinstance(rm_extra_sit, dict) and len(rm_extra_sit) > 0:
        for rm_code, extra_qty in rm_extra_sit.items():
            rm_code_str = str(rm_code)
            if rm_code_str in sit_simulated["RM CMMF"].values:
                sel = sit_simulated["RM CMMF"] == rm_code_str
                sit_simulated.loc[sel, "SIT Quantity"] = pd.to_numeric(
                    sit_simulated.loc[sel, "SIT Quantity"], errors="coerce"
                ).fillna(0) + extra_qty
            else:
                sit_simulated = pd.concat(
                    [sit_simulated, pd.DataFrame([{"RM CMMF": rm_code_str, "SIT Quantity": extra_qty}])],
                    ignore_index=True
                )

    # ---- Reload BOM and merge simulated SIT ----
    bom_data = load_and_clean_bom(demand_planning_file)
    bom_data["RM CMMF"] = bom_data["RM CMMF"].astype(str)

    sit_merged = pd.merge(bom_data, sit_simulated, on="RM CMMF", how="left")
    sit_merged["SIT Quantity"] = sit_merged["SIT Quantity"].fillna(0)
    sit_merged["Total Stock"] = sit_merged["RM In Stock"] + sit_merged["SIT Quantity"]

    # ---- ✅ Apply ALL FG target changes ----
    simulated_target_plan = (st.session_state.target_plan.copy()
                             if "target_plan" in st.session_state else {})
    if "fg_new_targets" in locals() and isinstance(fg_new_targets, dict) and len(fg_new_targets) > 0:
        for fg_code, new_target in fg_new_targets.items():
            simulated_target_plan[fg_code] = int(new_target)

    # ---- Re-run analysis WITH target using simulated data ----
    sit_suffix = "_sim"
    # Choose allocation method for simulation based on current selection
    if allocation_method == "Equally Distribution":
        sit_allocated = allocate_equal(sit_merged.copy(), "Total Stock", sit_suffix)
    elif allocation_method == "Production Plan Based (Target)":
        sit_allocated = allocate_by_target(
            sit_merged.copy(), "Total Stock", simulated_target_plan, sit_suffix
        )
    else:
        sit_allocated = allocate_by_price(sit_merged.copy(), "Total Stock", sit_suffix)

    family_summary_sim, bottleneck_analysis_sim, _ = run_bottleneck_analysis_with_target(
        sit_allocated,
        "Total Stock",
        f"Allocated Stock{sit_suffix}",
        simulated_target_plan,
        sit_suffix,
    )
    fg_summary_sim = create_fg_summary_view(
        bottleneck_analysis_sim,
        f"Allocated Stock{sit_suffix}",
        sit_suffix,
    )

    # ---- Show results ----
    st.subheader("📊 Simulated FG View Summary")
    st.dataframe(fg_summary_sim)

    st.subheader("⚠️ Simulated Bottleneck Analysis")
    st.dataframe(bottleneck_analysis_sim)

    # Compare to baseline if available
    base_sum = None
    try:
        base_df = st.session_state.results.get('with_target', {}).get('sit', {}).get('fg')
        if base_df is not None and not base_df.empty:
            base_sum = base_df["QTY to be produced"].sum()
        else:
            base_df2 = st.session_state.results.get('without_target', {}).get('sit', {}).get('fg')
            if base_df2 is not None and not base_df2.empty:
                base_sum = base_df2["QTY to be produced"].sum()
    except Exception:
        pass

    if base_sum is not None:
        delta_units = fg_summary_sim["QTY to be produced"].sum() - base_sum
        st.metric("🔁 Production Difference vs Baseline", f"{delta_units:+,.0f} Units")

    # ---- 📥 Download Options (include scenario changes) ----
    st.subheader("📥 Download Options (What-If)")

    # Build scenario change tables
    targets_df = pd.DataFrame(columns=["F.G. CMMF", "Old Target", "New Target", "Delta Target"])
    if "fg_new_targets" in locals() and isinstance(fg_new_targets, dict) and len(fg_new_targets) > 0:
        targets_df = pd.DataFrame(
            [
                {
                    "F.G. CMMF": str(fg),
                    "Old Target": int(st.session_state.target_plan.get(fg, 0)),
                    "New Target": int(new),
                    "Delta Target": int(new) - int(st.session_state.target_plan.get(fg, 0))
                }
                for fg, new in fg_new_targets.items()
            ]
        )

    rm_changes_df = pd.DataFrame(columns=["RM CMMF", "Extra SIT Quantity"])
    if "rm_extra_sit" in locals() and isinstance(rm_extra_sit, dict) and len(rm_extra_sit) > 0:
        rm_changes_df = pd.DataFrame(
            [{"RM CMMF": str(rm), "Extra SIT Quantity": int(qty)} for rm, qty in rm_extra_sit.items()]
        )

    # Create a single Excel with results + scenario sheets
    try:
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            # Results
            # Results (hide _sim requirement/surplus columns)
            _sim_hide = ["Required RM for Max Production_sim", "RM Surplus after Production_sim"]

            fg_summary_sim.to_excel(writer, sheet_name="Amiras_FG_View", index=False)
            bottleneck_analysis_sim.drop(columns=[c for c in _sim_hide if c in bottleneck_analysis_sim.columns],
                                        errors="ignore").to_excel(
                writer, sheet_name="Bottleneck_Analysis_SIT", index=False
            )
            family_summary_sim.to_excel(writer, sheet_name="Family_Summary", index=False)
            # Scenario inputs
            targets_df.to_excel(writer, sheet_name="Scenario_Target_Changes", index=False)
            rm_changes_df.to_excel(writer, sheet_name="Scenario_RM_SIT_Changes", index=False)

        st.download_button(
            label="📥 Download What-If Results (Excel)",
            data=output_buffer.getvalue(),
            file_name=f"FGtoRMAT_WhatIf_WithTarget_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Optional: separate quick CSVs for the changes
        if not targets_df.empty:
            st.download_button(
                label="⬇️ Download Target Changes (CSV)",
                data=targets_df.to_csv(index=False).encode("utf-8"),
                file_name=f"Scenario_Target_Changes_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )
        if not rm_changes_df.empty:
            st.download_button(
                label="⬇️ Download RM SIT Changes (CSV)",
                data=rm_changes_df.to_csv(index=False).encode("utf-8"),
                file_name=f"Scenario_RM_SIT_Changes_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )
    except Exception as e:
        st.warning(f"Could not build What-If download: {e}")

    st.success("✅ Simulation complete. Scroll up to compare with original results or run another scenario.")


# ==============================================================================
# Instructions (only show if no analysis completed)
if not st.session_state.get("analysis_completed", False):
    st.info("Follow these steps to generate your reports and visuals.")

    st.markdown("""
### 📋 Quick Start

**1) Upload files (Sidebar)**
- **Demand Planning (BOM)** — *Required*
- **SIT files** — *Optional* (In-Transit inventory)

**2) Select Analysis Type**
- **With Target** · **Without Target** · **Both**

**3) Date Range Filter for SIT**
- Use **Range start month/year** and **Range end month/year** to define the inclusive period for SIT data.  
- If any of the start or end values are **All**, no date filtering will be applied (all months/years are included).  
- For example, selecting **July 2025** as start and **September 2025** as end will include records for July, August, and September of 2025.

**4) Run Analysis**
- Click **Run Analysis** to process and build outputs

**5) Download Results**
- **FGtoRMAT(Imported+Local)Withtarget.xlsx** — with production targets  
- **FGtoRMAT(Imported+Local)Withouttarget.xlsx** — without targets  
- Available for **Original Data** and **With SIT Data**
""")

    with st.expander("📑 Required columns by file (please match names/case)", expanded=True):
        st.markdown("""
**Demand Planning (BOM) — Required**
- `F.G. CMMF` *(text)*  
- `Description` *(text)*  
- `Family` *(text)*  
- `RM CMMF` *(text)*  
- `RM Description` *(text)*  
- `Qty / Unit` *(number)*  
- `RM In Stock` *(number)*  

**BOM — Recommended (enables richer financials & KPIs)**
- `RM Value` *(number; unit value of RM)*  
- `Price` *(number; FG unit price)*  
- `Market` *(text)*  
- `Planned Production` or `Target` *(number; optional)*

**SIT Files — Minimum**
- `Item No.` **or** `RM CMMF` *(text; RM code)*  
- `Quantity` **or** `Qty` *(number; in-transit qty)*

**SIT Files — For date filtering (any one is enough)**
- One of: `ETD`, `ATD`, `ETA`, `ATA`, `Shipping Date`, `Expected_ETD`, `DocDate`, `Posting Date` *(date)*

**Notes**
- Treat **`F.G. CMMF`** and **`RM CMMF`** as **Text** in Excel to avoid scientific/float formatting (e.g., `43000075`, not `4.3000075E+7`).  
- Put column headers on the **first row**, no merged cells.  
- If a file won’t parse, re-save it as **.xlsx** (Excel Workbook).
""")

    with st.expander("🔍 What’s inside (features)"):
        st.markdown("""
- **Exact logic**: Mirrors the original scripts’ algorithms  
- **Dynamic processing**: Uses your uploaded files  
- **Interactive visuals**: Toggle Original vs SIT  
- **Professional Excel output**: Structured, formatted sheets  
- **Comprehensive analysis**: KPIs, production, bottlenecks, financials  
- **Smart SIT date filtering**: Supports “All” month/year  
- **Session persistence**: Results stay until cleared  
- **Progress & errors**: Clear status and helpful messages
""")

    with st.expander("🛠️ Tips & troubleshooting"):
        st.markdown("""
- Verify BOM columns exactly match the list above  
- If outputs look empty, recheck **date filters** and **analysis type**  
- After updating files, click **Run Analysis** again to refresh results
""")

# Footer
st.markdown("---")
st.markdown("*Production Planning Dashboard - Enhanced with Professional Excel Downloads & Improved User Experience*")
